import logging
import os.path
from io import BytesIO

import sqlalchemy
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile, BufferedInputFile, CallbackQuery
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from telegraph import Telegraph

from handlers.start import choose_language
from keyboards.default.button import settings_kb, study_menu
from utils.db.db_sqlalchemy import Category, SubCategory, Quiz, Option, session, UserAnswer
from utils.helper.user_helper import show_statistics, is_admin, all_quizzes

menu_router = Router()

tg = Telegraph()


@menu_router.message(F.text == '📚 Darsliklar')
async def educational_lesson(message: Message):
    tg.create_account(short_name='1337')

    response = tg.create_page(
        'Assalomu alaykum!',
        html_content='<p>📚 Darsliklar</p>'
    )
    await message.answer(text=f"<a href=\"{response['url']}\">📚 Darsliklar</a>")


@menu_router.message(F.text == '👨‍🏫 Kurslar')
async def educational_course(message: Message):
    tg.create_account(short_name='1337')

    response = tg.create_page(
        'Assalomu alaykum!',
        html_content='<p>👨‍🏫 Kurslarimiz</p>'
    )
    await message.answer(text=f"<a href=\"{response['url']}\">👨‍🏫 Kurslar</a>")


@menu_router.message(F.text == '📊 Statistika')
async def educational_statistics(message: Message):
    tg.create_account(short_name='1337')
    results = show_statistics()

    html_content = ("<p><b>📊 Umumiy statistika natijalari</b></p>"
                    "<p><b>Quyidagi jadvalda foydalanuvchilarning testdagi ishtiroki ko'rsatilgan:</b></p>")

    for index, (fullname, correct, total, percent) in enumerate(results, start=1):
        html_content += (
            f"<p style='border:1px solid #ccc; padding:10px; margin:10px 0;'>"
            f"<b>👤 {index} {fullname}</b>"
            f"<b> - {percent}%</b><br>"
            f"✅ To'g'ri javoblar: <b>{correct}</b><br>"
            f"📋 Jami javoblar: <b>{total}</b>"
            f"</p>"
        )

    response = tg.create_page(
        title="Statistika",
        html_content=html_content
    )

    await message.answer(
        text=f"<a href=\"{response['url']}\">👨‍🏫 Statistika</a>",
        parse_mode="HTML"
    )


@menu_router.message(F.text == "🌐 Tilni o'zgartirish")
async def again_choose_language(message: Message):
    await choose_language(message)


@menu_router.message(F.text == '⬅️ Ortga')
async def basic_menu(message: Message):
    chat_id = message.chat.id
    await message.answer('Quyidagilardan birini tanang', reply_markup=study_menu(chat_id))


@menu_router.message(F.text == '⚙️ Sozlamalar')
async def settings(message: Message):
    await message.answer(text='Quyidagi birini tanlang.', reply_markup=settings_kb())


@menu_router.message(F.text == 'Savollarni yuklab olish')
async def download_xls(message: Message):
    wb = Workbook()
    ws = wb.active
    ws.title = 'savollar'
    ws.append(['Category', 'Subcategory', 'Quiz', 'A', 'B', 'C', 'D'])
    for quiz in all_quizzes():
        category_name = quiz.subcategory.category.name
        subcategory_name = quiz.subcategory.name
        text = quiz.text
        a = next(opt.text for opt in quiz.options if opt.is_correct)
        b, c, d = [opt.text for opt in quiz.options if not opt.is_correct]
        ws.append([category_name, subcategory_name, text, a, b, c, d])
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = BufferedInputFile(file.read(), filename='all_quizzes.xlsx')
    await message.answer_document(document=document, caption='Bo`limlarga oid barcha savollar yuklandi.')


@menu_router.message(F.text == "Shablon")
async def send_template(message: Message):
    wb = Workbook()
    ws = wb.active
    ws.append(["Category", "Subcategory", "Quiz", "A_true", "B", "C", "D"])
    ws.append(["Math", "Algebra", "2 + 2 = ?", "4", "3", "5", "2"])

    file = BytesIO()
    wb.save(file)
    file.seek(0)
    input_file = BufferedInputFile(file.read(), filename="quiz_template.xlsx")
    await message.answer_document(document=input_file, caption='Bo`lim va savollarni qo`shish na`muna')


@menu_router.message(F.document.file_name.endswith('.xlsx'))
async def handle_xlsx_upload(message: Message):
    file = await message.bot.get_file(message.document.file_id)
    file_bytes = await message.bot.download_file(file.file_path)
    file_content = BytesIO(file_bytes.read())

    wb = load_workbook(file_content)
    sheet = wb.active

    stats = {"categories": 0, "subcategories": 0, "quizzes": 0}
    required_columns = 7

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < required_columns or any(row[i] is None for i in range(3)):
            await message.answer(f"{idx}-qatordagi ma’lumotlar to‘liq emas. O‘tkazib yuborildi.")
            continue

        category_name, subcategory_name, quiz_text, a_true, b, c, d = row

        # Category
        category = session.query(Category).filter_by(name=category_name).first()
        if not category:
            category = Category(name=category_name)
            session.add(category)
            session.flush()
            stats["categories"] += 1

        # SubCategory
        subcategory = session.query(SubCategory).filter_by(name=subcategory_name, category_id=category.id).first()
        if not subcategory:
            subcategory = SubCategory(name=subcategory_name, category_id=category.id)
            session.add(subcategory)
            session.flush()
            stats["subcategories"] += 1

        # Quiz dublikatni tekshirish
        existing_quiz = session.query(Quiz).filter_by(text=quiz_text, subcategory_id=subcategory.id).first()
        if existing_quiz:
            continue

        # Quiz
        quiz = Quiz(subcategory_id=subcategory.id, text=quiz_text)
        session.add(quiz)
        session.flush()
        stats["quizzes"] += 1

        # Options
        options = [
            (a_true, True),
            (b, False),
            (c, False),
            (d, False),
        ]
        for text, is_correct in options:
            if text:
                session.add(Option(quiz_id=quiz.id, text=text, is_correct=is_correct))

    session.commit()

    await message.answer(
        "✅ Ma’lumotlar bazaga muvaffaqiyatli yozildi:\n"
        f"📂 Kategoriyalar: {stats['categories']}\n"
        f"📁 Subkategoriyalar: {stats['subcategories']}\n"
        f"❓ Savollar: {stats['quizzes']}"
    )


@menu_router.message(F.text == 'O`chirish')
async def delete_note(message: Message):
    await message.answer(
        "❗️ Diqqat: O'chirish qoidasi:\n\n"
        "To‘g‘ri foydalanish:\n"
        "/find_quiz kalit so‘z\n"
        "/find_category kalit so‘z\n"
        "/find_subcategory kalit so‘z\n\n"
        "Misol:\n"
        "Copy: <code>/find_quiz suv</code>\n"
        "Copy: <code>/find_category texnologiya</code>\n"
        "Copy: <code>/find_subcategory daryo</code>"
    )


@menu_router.message(F.text.startswith('/find'))
async def find_quiz_handler(message: Message):
    if not is_admin(message.from_user.id):
        return await message.answer("❌ Sizda ruxsat yo'q.")

    keywords = message.text.split(maxsplit=1)
    if not any(cmd in keywords for cmd in ['/find_quiz', '/find_category', '/find_subcategory']):
        return await message.answer(
            "❗️ Noto‘g‘ri format.\n\n"
            "To‘g‘ri foydalanish:\n"
            "/find_quiz kalit so‘z\n"
            "/find_category kalit so‘z\n"
            "/find_subcategory kalit so‘z\n\n"
            "Misol:\n"
            "Copy: <code>/find_quiz suv</code>\n"
            "Copy: <code>/find_category texnologiya</code>\n"
            "Copy: <code>/find_subcategory daryo</code>"
        )

    search_term = keywords[1]
    prefix = keywords[0]
    if prefix == '/find_quiz':
        quizzes = session.query(Quiz).filter(Quiz.text.ilike(f"%{search_term}%")).all()
    elif prefix == '/find_category':
        quizzes = session.query(Category).filter(Category.name.ilike(f"%{search_term}%")).all()
    elif prefix == '/find_subcategory':
        quizzes = session.query(SubCategory).filter(SubCategory.name.ilike(f"%{search_term}%")).all()
    else:
        quizzes = None

    if not quizzes:
        return await message.answer("🔍 Hech qanday savol topilmadi.")

    builder = InlineKeyboardBuilder()
    prefix = prefix.split('_')[-1]
    for quiz in quizzes:
        builder.button(
            text=f"{quiz}...",
            callback_data=f"delete_{prefix}:{quiz.id}"
        )
    builder.adjust(1)

    await message.answer("🧾 Topilgan savollar:", reply_markup=builder.as_markup())


@menu_router.callback_query(F.data.startswith("delete"))
async def delete_quiz_callback(callback: CallbackQuery):
    logging.info(callback.data)
    if not is_admin(callback.from_user.id):
        return await callback.answer("❌ Ruxsat yo'q", show_alert=True)

    split_texts = callback.data.split(":")
    object_id = int(split_texts[1])
    prefix = split_texts[0].split('_')[1]
    try:
        if prefix == 'quiz':
            quiz = session.get(Quiz, object_id)
            if not quiz:
                return await callback.answer("❌ Savol topilmadi", show_alert=True)
            for opt in quiz.options:
                try:
                    session.delete(opt)
                except Exception as e:
                    logging.warning(f'{e}')
                    ua = session.query(UserAnswer).filter(UserAnswer.option_id == opt.id).first()
                    session.delete(ua)
            session.commit()
            session.delete(quiz)
        elif prefix == 'subcategory':
            subcategory = session.get(SubCategory, object_id)
            if not subcategory:
                return await callback.answer("❌ Subkategroiya topilmadi", show_alert=True)
            for quiz in subcategory.quizzes:
                for opt in quiz.options:
                    session.delete(opt)
                session.delete(quiz)
            session.delete(subcategory)
        elif prefix == 'category':
            category = session.get(Category, object_id)
            if not category:
                return await callback.answer("❌ Kategoriya topilmadi", show_alert=True)
            for sub in category.subcategories:
                for quiz in sub.quizzes:
                    for opt in quiz.options:
                        session.delete(opt)
                    session.delete(quiz)
                session.delete(sub)
            session.delete(category)
        session.commit()
        await callback.message.edit_text(f"🗑 Obyekt o‘chirildi:")
    except sqlalchemy.exc.IntegrityError as e:
        print(e)
        await callback.answer(f'⚠️ IntegrityError:')
    except Exception as e:
        print(e)
        await callback.answer(f'⚠️ Exception:')



